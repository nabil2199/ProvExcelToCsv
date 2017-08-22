#! python3

import openpyxl, os, sys

#wbpath=sys.argv[1]
#currentworkdirectory=os.getcwd()

#print(wbpath)
#workbook=openpyxl.load_workbook(wbpath, data_only=True)
#workbook.get_sheet_names()

#load and initialise destination csv file
csvfile = open('C:\\Users\\NXLX8474\\Desktop\\todo20170824\\lotb5\\lotb5test.csv', 'w')
csvfile.write('upn,FirstName,LastName,SipAddress,EmailAddress,TelURI,Extension,VoicePolicy,Tactical,VoiceMail,Desk,DisplayNumber\n')


workbook = openpyxl.load_workbook('C:\\Users\\NXLX8474\\Desktop\\todo20170824\\lotb5\\lotb5.xlsm', data_only=True)
workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name('Lot A - User Information')
sheetsize = sheet.max_row

#def getUserInfoSheetSize(workbookfile):
#    size=1
#    cellcoordinate="F"+"size"
#    cell=sheet[cellcoordinate]
#    while (cell.value!=0 and cell.value!=None):
#        size = size + 1
#    return size

def loadColumn(columnindex):
    valuelist = []
    i=1
    while (i < sheetsize):
        cellcoordinate=columnindex + str(i)
        cell = sheet[cellcoordinate]
        valuelist.insert(i,cell.value)
        i=i+1
    return valuelist

def loadUPN():
    upnlist = loadColumn("E")
    return upnlist

def loadSIP():
    upnlist = loadColumn("F")
    return upnlist

def loadFName():
    fnamelist = loadColumn("B")
    return fnamelist

def loadLName():
    lnamelist = loadColumn("C")
    return lnamelist

def loadEMail():
    emaillist = loadColumn("H")
    return emaillist

def loadLTelUri():
    telurilist = loadColumn("I")
    return telurilist

def loadExtension():
    extensionlist = loadColumn("J")
    return extensionlist

def loadVoicePolicy():
    voicepolicylist = loadColumn("M")
    return voicepolicylist

#def loadTactical():
#    tacticallist = loadColumn("S")
#    return tacticallist

def loadTactical():
    valuelist = []
    i=1
    while (i < sheetsize):
        cellcoordinate='S' + str(i)
        cell = sheet[cellcoordinate]
        if cell.value == 'Y':
            valuelist.insert(i,'True')
        elif cell.value == 'N':
            valuelist.insert(i, 'False')
        else :
            valuelist.insert(i,cell.value)
        i=i+1
    return valuelist

def loadVoiceMail():
    voicemaillist = loadColumn("V")
    return voicemaillist

def loadDesk():
    desklist = loadColumn("AD")
    return desklist

def loadDisplayNumber():
    displaynumberlist = loadColumn("R")
    return displaynumberlist

bupnlist = loadUPN()
fnamelist = loadFName()
lnamelist = loadLName()
siplist = loadSIP()
emaillist = loadEMail()
telurilist = loadLTelUri()
extensionlist = loadExtension()
voicepolicylist = loadVoicePolicy()
tacticallist = loadTactical()
voicemaillist = loadVoiceMail()
desklist = loadDesk()
displaynumberlist = loadDisplayNumber()

for i in range(1,sheetsize-1):
    if bupnlist[i] != None and bupnlist[i] != 0 and desklist[i] == None:
        line = str(bupnlist[i]) + ',' + str(fnamelist[i]) + ',' + str(lnamelist[i]) + ',' + str(siplist[i]) + ',' + str(emaillist[i]) + ',' + str(telurilist[i]) + ',' + str(extensionlist[i]) + ',' + str(voicepolicylist[i]) + ',' + str(tacticallist[i]) + ',' + str(voicemaillist[i]) + ',' +  ',' + str(displaynumberlist[i] + '\n')
        csvfile.write(line)
        print(i)
        print(line)
    elif bupnlist[i] != None and bupnlist[i] != 0 and desklist[i] != None:
        line = str(bupnlist[i]) + ',' + str(fnamelist[i]) + ',' + str(lnamelist[i]) + ',' + str(siplist[i]) + ',' + str(emaillist[i]) + ',' + str(telurilist[i]) + ',' + str(extensionlist[i]) + ',' + str(voicepolicylist[i]) + ',' + str(tacticallist[i]) + ',' + str(voicemaillist[i]) + ',' + str(desklist[i]) + ',' + str(displaynumberlist[i] + '\n')
        csvfile.write(line)
        print(i)
        print(line)

csvfile.close()